"""
This is the main module for this repository
"""

import shutil
import json
from pathlib import Path
from . import keller, pathfolio, oh_my_ontology
from openpyxl import load_workbook

class Identifiers:
    """
    Class object help with the identification of space, project and experiment in openBIS.
    """
    def __init__(self, space_code: str, project_code: str, experiment_code: str) -> None:
        self.space_code = space_code
        self.project_code = project_code
        self.experiment_code = experiment_code

    @property
    def space_identifier(self) -> str:
        return self.space_code

    @property
    def project_identifier(self) -> str:
        return f"/{self.space_identifier}/{self.project_code}"

    @property
    def experiment_identifier(self) -> str:
        return f"{self.project_identifier}/{self.experiment_code}"
    
class Dataset:
    """
    Class made to facilitate dataset upload
    """
    def __init__(self, openbis_instance, ident: Identifiers, dataset_type=None, upload_data=None) -> None:
        self.ob = openbis_instance
        self.ident = ident
        self.type = dataset_type
        self.data = upload_data
        self.experiment = self.ident.experiment_identifier.upper()  # Use the provided Identifiers instance

    def upload_dataset(self):
        """
        Upload the dataset to the openBIS
        """
        self.ob.new_dataset(type=self.type, experiment=self.experiment, file=self.data).save()


def push_exp(
        dir_pat: str,
        dir_folder: str,
        user_mapping: dict = None,
        dict_mapping: dict = pathfolio.metadata_mapping,
        space_code: str = 'TEST_SPACE_PYBIS',
        project_code: str = 'TEST_UPLOAD',
        experiment_type: str = 'Battery_Premise2',
) -> None:
    """
    Pushes experimental data and metadata from a local folder to an openBIS instance.

    Args:
        dir_pat (str): Path to the openBIS PAT file (personal access token).
        dir_folder (str): Path to the directory containing the experimental data files.
        user_mapping (dict, optional): A dictionary mapping short name codes to full names.
        dict_mapping (dict, optional): A mapping dictionary defining openBIS codes and JSON paths for metadata extraction. Defaults to `pathfolio.metadata_mapping`.
        space_code (str, optional): The openBIS space code where the experiment will be created. Defaults to 'TEST_SPACE_PYBIS'.
        project_code (str, optional): The openBIS project code where the experiment will be created. Defaults to 'TEST_UPLOAD'.
        experiment_type (str, optional): The type of experiment to be created in openBIS. Defaults to 'Battery_Premise2'.

    Raises:
        ValueError: If there is not exactly one JSON file in the specified folder.
        ValueError: If the JSON file name does not follow the required naming convention.
        ValueError: If there is not exactly one raw HDF5 file in the specified folder.

    Returns:
        None
    """
    dir_folder = Path(dir_folder)
    ob = keller.get_openbis_obj(dir_pat)

    list_json = [file for file in dir_folder.iterdir() if file.suffix == ".json" and not file.stem.startswith('ontologized')]
    if len(list_json) != 1:
        raise ValueError("There should be exactly one json file in the folder")

    name_json = list_json[0].name
    dir_json = dir_folder / name_json

    if len(dir_json.name.split('.')) != 3:
        raise ValueError("Not recognized json file name. The recognized file name is cycle.experiment_code.json")

    exp_name = dir_json.stem.split('.')[1]  # Extract the experiment name from the json file name
    ident = Identifiers(space_code, project_code, experiment_code=exp_name)

    # Create new experiment in the predefined space and project.
    exp = ob.new_experiment(code=ident.experiment_code, type=experiment_type, project=ident.project_identifier)

    # Iterate through list of metadata from json file and upload them to the experiment.
    with Path(dir_json).open() as f:
        sample_metadata = json.load(f)["metadata"]["sample_data"]

    for item in dict_mapping:
        try:
            key = item['metadata']
            openbis_code = item['openbis_code']
            print(f'Uploading metadata for {openbis_code} from {key}')
            exp.p[openbis_code] = sample_metadata.get(key)
            exp.save()
        except Exception as e:
            print(f'Error uploading metadata for {openbis_code} from {key}')
            print(f'The error message is: {e} \n')
            continue

    # Upload the dataset

    # Analyzed data
    ds_analyzed_json = Dataset(ob, ident=ident)
    ds_analyzed_json.type = 'premise_cucumber_analyzed_battery_data'
    ds_analyzed_json.data = dir_json
    ds_analyzed_json.upload_dataset()

    # Raw data
    list_raw_data = [file for file in dir_folder.iterdir() if file.suffix == ".h5" and file.stem.split('.')[1] == exp_name]
    if len(list_raw_data) != 1:
        raise ValueError("There should be exactly one raw_h5 file in the folder")

    dir_raw_json = list_raw_data[0]
    ds_raw_data = Dataset(ob, ident=ident)
    ds_raw_data.type = 'premise_cucumber_raw_battery_data'
    ds_raw_data.data = dir_raw_json
    ds_raw_data.upload_dataset()

    # Create the automated_extract_metadata.xlsx file
    oh_my_ontology.gen_metadata_xlsx(dir_json, user_mapping=user_mapping)
    source_file = dir_folder / f"{exp_name}_automated_extract_metadata.xlsx"
    dest_file = dir_folder / f"{exp_name}_merged_metadata.xlsx"
    print(f"Copying {source_file} to {dest_file}")
    shutil.copy(source_file, dest_file)

    # Check if there is already a custom Excel file for the experiment. If so, create JSON-LD from it.
    custom_metadata_files = [file for file in dir_folder.iterdir() if file.name.endswith('custom_metadata.xlsx')]
    if custom_metadata_files:
        custom_metadata = custom_metadata_files[0]

        # Load both Excel files
        merged_wb = load_workbook(dest_file)
        custom_wb = load_workbook(custom_metadata)

        # Select the "Schema" sheet from both workbooks
        merged_sheet = merged_wb["Schema"]
        custom_sheet = custom_wb["Schema"]

        # Find the "Value" column index in the "Schema" sheet
        header_row = 1  # Assuming headers are in the first row
        value_column_index = None

        for col in range(1, custom_sheet.max_column + 1):
            if custom_sheet.cell(row=header_row, column=col).value == "Value":
                value_column_index = col
                break

        if value_column_index is None:
            raise ValueError("Column 'Value' not found in the 'Schema' sheet.")

        # Loop through rows in the "Value" column of the custom metadata
        for row in range(header_row + 1, custom_sheet.max_row + 1):  # Skip the header row
            custom_value = custom_sheet.cell(row=row, column=value_column_index).value
            if custom_value:  # Skip if the cell is empty or None
                # Write the custom value into the corresponding row of the merged metadata
                merged_sheet.cell(row=row, column=value_column_index).value = custom_value

        # Save the updated merged metadata workbook
        merged_wb.save(dest_file)

    # Upload the metadata Excel file to the openBIS
    dir_metadata_excel = dir_folder / f"{exp_name}_merged_metadata.xlsx"
    ds_metadata_excel = Dataset(ob, ident=ident)
    ds_metadata_excel.type = 'premise_excel_for_ontology'
    ds_metadata_excel.data = dir_metadata_excel
    ds_metadata_excel.upload_dataset()

    # Generate the ontologized JSON-LD file
    jsonld_filename = f"ontologized_{exp_name}.json"
    dir_xlsx = dir_folder / f"{exp_name}_merged_metadata.xlsx"
    oh_my_ontology.gen_jsonld(dir_xlsx, jsonld_filename)

    # Upload the ontologized JSON-LD file to the openBIS
    dir_jsonld = dir_folder / jsonld_filename
    ds_jsonld = Dataset(ob, ident=ident)
    ds_jsonld.type = 'premise_jsonld'
    ds_jsonld.data = dir_jsonld
    ds_jsonld.upload_dataset()
