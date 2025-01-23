"""Functions to handle ontology xlsx file generation and upload."""

import json
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from . import pathfolio, simon_simulator


def update_metadata_value(file_path: str, metadata: str, input_value: str, sheet_name: str = "Schema"):
    """Update the value of a specified metadata key in an Excel sheet.

    Args:
        file_path (str): Path to the Excel file.
        metadata (str): The metadata key to search for.
        input_value (str): The new value to set.
        sheet_name (str): Name of the sheet to search in (default is "Schema").

    """
    try:
        # Load the workbook and select the specified sheet
        workbook = load_workbook(file_path)
        if sheet_name not in workbook.sheetnames:
            print(f"Sheet '{sheet_name}' not found in the workbook.")
            return

        sheet = workbook[sheet_name]

        # Find the row with the specified metadata and update the corresponding value
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=2):
            metadata_cell, value_cell = row
            if metadata_cell.value == metadata:
                value_cell.value = input_value
                print(f"Updated metadata '{metadata}' with value '{input_value}'.")
                break
        else:
            print(f"Metadata '{metadata}' not found in sheet '{sheet_name}'.")

        # Save the changes to the Excel file
        workbook.save(file_path)

    except Exception as e:
        print(f"An error occurred: {e}")

def gen_metadata_xlsx(
        dir_json: str,
        user_mapping: dict = None,
        dir_template: str = r"K:\Aurora\nukorn_PREMISE_space\Battinfo_template.xlsx",
    ) -> None:
    r"""Generate a metadata Excel file for a specific experiment based on a template.

    After generating this metadata file, the script will automatically extract the metadata information form the
    analyzed json file and automatcailly update the metadata values in the new Excel file.
    This Excel file will then be used to generate a corresponding ontologized JSON-LD file.

    Args:
        dir_json (str): The path to the analyzed JSON file.
        dir_template (str): The path to the template Excel file. Defaults to
                            'K:\\Aurora\\nukorn_PREMISE_space\\Battinfo_template.xlsx'.
        user_mapping (dict, optional): A dictionary mapping user short names to full names.

    Returns:
        None: Creates a new Excel file in the backup directory with the experiment name as part of the file name.

    """
    dir_json = Path(dir_json)
    # Extract the experiment name from the analyzed JSON file path
    experiment_name = dir_json.stem.split(".")[1]
    # Get the name of the new xlsx file
    new_xlsx_name = f"{experiment_name}_automated_extract_metadata.xlsx"
    dir_new_xlsx = dir_json.parent / new_xlsx_name
    # Copy the template file
    shutil.copy(dir_template, dir_new_xlsx)

    # Update the experiment name in the new Excel file
    dict_metadata = curate_metadata_dict(dir_json, user_mapping=user_mapping)
    for key, value in dict_metadata.items():
        update_metadata_value(dir_new_xlsx, key, value)


def curate_metadata_dict(dir_json: str, user_mapping: dict | None = None) -> dict[str, str]:
    """Generate a metadata dictionary by extracting relevant information from a JSON file.

    This metadata is used to populate an Excel file that will later generate an
    ontologized JSON-LD file for further use.

    Args:
        dir_json (str): The file path to the JSON file that contains the analyzed metadata.
        user_mapping (dict, optional): A dictionary mapping user short names to full names.

    Returns:
        Dict[str, str]: A dictionary containing metadata items as keys and their corresponding values.

    Raises:
        ValueError: If an assembly date cannot be extracted.

    """
    dict_metadata = {}

    with Path(dir_json).open() as f:
        sample_metadata = json.load(f)["metadata"]["sample_data"]

    #Extract metadata from the analyzed json file.
    for key, value in pathfolio.dict_excel_to_json.items():
        dict_metadata[key] = sample_metadata.get(value)

    #Extracting operator name
    user_short_name = dict_metadata["Cell ID"].split("_")[1]
    user_full_name = user_mapping.get(user_short_name, user_short_name) if user_mapping else user_short_name

    dict_metadata["Scientist/technician/operator"] = user_full_name


    #Extracting the date of the experiment
    try:
        parsed_date = datetime.strptime(
            sample_metadata["Timestamp step 10"],
            "%Y-%m-%d %H:%M:%S",
        )
        dict_metadata["Date of cell assembly"] = parsed_date.strftime("%d/%m/%Y")
    except:
        msg = f"Could not extract datetime from sample {dict_metadata.get('Cell ID', 'Unknown')}"
        raise ValueError(msg)
    return dict_metadata

def gen_jsonld(dir_xlsx: str,  jsonld_filename: str) -> None:
    """Generate a JSON-LD file from a metadata Excel file.

    Args:
        dir_xlsx (str): The path to the metadata Excel file.
        jsonld_filename (str): The name of the JSON-LD file.

    Returns:
        None: Creates a new JSON-LD file in the specified directory.

    """
    dir_xlsx = Path(dir_xlsx)
    json_ld_output = simon_simulator.convert_excel_to_jsonld(dir_xlsx)
    jsonld_str = json.dumps(json_ld_output, indent=4)
    jsonld_filepath = dir_xlsx.parent/jsonld_filename

    with open(jsonld_filepath, "w") as f:
        f.write(jsonld_str)
